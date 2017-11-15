from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from inspection.models import shelf, ShelfImport
from openpyxl.compat import range

# excel handling

# https://www.testwo.com/blog/7269
def import_shelf(self, request, obj, change):

    wb = load_workbook(filename=obj.shelf_import_file.path)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    headers = ['type', 'warehouse', 'compartment', 'warehouse_channel','group','number','is_gradient_measurement_mandatory']
    lists = []
    print ws.max_row
    for row in range(2, ws.max_row+1):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    sqllist = []
    for cell in lists:
        sql = shelf()
        for header in headers:
            setattr(sql,header,cell[header])           
        if not sql.is_exist():
            sqllist.append(sql)
    shelf.objects.bulk_create(sqllist)    